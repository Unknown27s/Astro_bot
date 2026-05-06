"""Simple CSV/Excel parser for student data and marks."""
import csv
import io
from pathlib import Path

def parse_students_csv(content: bytes, file_ext: str) -> list[dict]:
    """Parse student master data from CSV/XLSX."""
    if file_ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).lower().strip() for h in rows[0]]
            students = []
            for row in rows[1:]:
                if any(row):
                    students.append(dict(zip(headers, row)))
            return students
        except Exception as e:
            raise ValueError(f"Failed to parse XLSX: {e}")
    else:
        try:
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            # Normalize headers
            normalized_rows = []
            for row in reader:
                normalized_row = {str(k).lower().strip(): v for k, v in row.items()}
                normalized_rows.append(normalized_row)
            return normalized_rows
        except Exception as e:
            raise ValueError(f"Failed to parse CSV: {e}")

def parse_marks_csv(content: bytes, file_ext: str) -> list[dict]:
    """Parse student marks data from CSV/XLSX."""
    if file_ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).lower().strip() for h in rows[0]]
            marks = []
            for row in rows[1:]:
                if any(row):
                    marks.append(dict(zip(headers, row)))
            return marks
        except Exception as e:
            raise ValueError(f"Failed to parse XLSX: {e}")
    else:
        try:
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            # Normalize headers
            normalized_rows = []
            for row in reader:
                normalized_row = {str(k).lower().strip(): v for k, v in row.items()}
                normalized_rows.append(normalized_row)
            return normalized_rows
        except Exception as e:
            raise ValueError(f"Failed to parse CSV: {e}")

def parse_unified_csv(content: bytes, file_ext: str) -> list[dict]:
    """Parse unified student+marks data from CSV/XLSX.
    Expected columns: roll_no, name, email, phone, department, semester,
                      subject_code, subject_name, subject_semester, 
                      internal_marks, external_marks, grade
    """
    if file_ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).lower().strip() for h in rows[0]]
            data = []
            for row in rows[1:]:
                if any(row):
                    data.append(dict(zip(headers, row)))
            return data
        except Exception as e:
            raise ValueError(f"Failed to parse XLSX: {e}")
    else:
        try:
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            # Normalize headers
            normalized_rows = []
            for row in reader:
                normalized_row = {str(k).lower().strip(): v for k, v in row.items()}
                normalized_rows.append(normalized_row)
            return normalized_rows
        except Exception as e:
            raise ValueError(f"Failed to parse CSV: {e}")
